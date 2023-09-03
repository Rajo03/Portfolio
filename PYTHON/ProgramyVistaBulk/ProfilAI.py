import xlsxwriter
import datetime
import pandas as pd
import os
from PIL import Image
import openpyxl

date_str = input("podaj date pierwszego posta '%Y-%m-%d %H:%M': ")
sciezka = input("Podaj scieżke:")
ilosc_postow = int(input("ile postow chcesz zaplanować?"))


#TODO:Pobieranie opisu z innego arkusza


#rename pliki i convert na jpg
path = f"S:\\ARCHIWUM\\The Ai World\\dowrzucenia\\{sciezka}"





# Set up the input and output file paths
def tworzenie_pliku_excel():
    xlsx_file_path = "C:\\PROGRAMOWANIE\\PYTHON\\GotoweVistaBulk\\profil_Ai_Pinterest.xlsx"



    date = datetime.datetime.strptime(date_str, "%Y-%m-%d %H:%M")

    # Create a new XLSX file and add a worksheet
    workbook = xlsxwriter.Workbook(xlsx_file_path)
    worksheet = workbook.add_worksheet()

        
        

    # Set the column headers
    header_row = ["message", "type", "link", "time"]
    worksheet.write_row(0, 0, header_row)

    # Set up the CSV reader and read the column data

    # Write the column data and other data to the worksheet
    for i in range(ilosc_postow):
        
        source_workbook = openpyxl.load_workbook("C:\\Users\\kuba\\Desktop\\Wszystko\\ExcelAutomatize\\Profil_AI_Canvabulk.xlsx")
        source_worksheet = source_workbook['innyarkusz']


        # Write the message column data
        message_cell = source_worksheet.cell(row=i + 2, column=14)  # Kolumna "L" to kolumna 12
        message = message_cell.value
        worksheet.write(i + 1, 0, message)
        
        # Write the type column data
        worksheet.write(i + 1, 1, "image")
        link_index = i * 8 + 1
        # Write the link column data
        link_path = f"http://hosting2303687.online.pro/ARCHIWUM/The Ai World/dowrzucenia/{sciezka}/{link_index}.jpg"
        worksheet.write(i + 1, 2, link_path)
        
        # Write the time column data
        
        
        worksheet.write(i + 1, 3, date.strftime("%Y-%m-%d %H:%M"))
        
        
        date += datetime.timedelta(days = 1)
        
        



    # Close the workbook
    workbook.close()



def convert_xlsx_to_csv():
    df = pd.read_excel("C:\\PROGRAMOWANIE\\PYTHON\\GotoweVistaBulk\\profil_Ai_Pinterest.xlsx")

    df.to_csv("C:\\PROGRAMOWANIE\\PYTHON\\GotoweVistaBulk\\profil_Ai_Pinterest.csv", index=False)

    exit()




tworzenie_pliku_excel()
convert_xlsx_to_csv()